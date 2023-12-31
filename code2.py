import cv2 as cv
import os
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing import image
from tensorflow.keras.optimizers import RMSprop
from tensorflow.keras.preprocessing.image import ImageDataGenerator
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


os.chdir("R:/Face/model")


#To ppredict the face
def predict_face(model,F_name):
    cam=cv.VideoCapture(0)
    result,imaage=cam.read()
    del cam
    face=cv.resize(imaage, (200,200))    
    X=image.img_to_array(face)
    X=np.expand_dims(X,axis=0)
    images=np.vstack([X])
    val=model.predict(images)
    val=np.ndarray.tolist(val)
    val=val[0]
    if val[0]==1:
        print('Rohith')
        return 'Rohith'
    elif val[1]==1:
        print('Saji')
        return 'Saji'
    elif val[2]==1:
        print('Syna')
        return 'Syna'
    elif val[3]==1:
        print(F_name)
        return F_name
    else:
        print('none') 
        return None
    


#To train new model with new face
def train_model():
    model = load_model('FR1.h5')
   
    train= ImageDataGenerator(rescale=1/255)
    validation=ImageDataGenerator(rescale=1/255)
    train_dataset=train.flow_from_directory("R:/Face/model/new-train/",
                                            target_size=(200,200),
                                            batch_size=20,
                                            class_mode='categorical')

    validation_dataset=validation.flow_from_directory("R:/Face/model/new-validation/",
                                            target_size=(200,200),
                                            batch_size=14,
                                            class_mode='categorical')
 
    model.compile(loss='categorical_crossentropy',optimizer=RMSprop(learning_rate=0.001),metrics=['accuracy'])

    model.fit(train_dataset,
                        steps_per_epoch=100,
                        epochs=10,
                        validation_data=validation_dataset,
                        validation_steps=100)
    return model
    


#To capture images
def capt_img(p_directory,F_name,n):
    cam=cv.VideoCapture(0)
    img_no=0
    while True:
        result,imaage=cam.read()
        img_no+=1
        face=cv.resize(imaage, (200,200))
        #face=cv.cvtColor(face,cv.COLOR_BGR2GRAY)
        path=p_directory+'/'+F_name+'/'+F_name+str(img_no)+'.jpg'
        cv.imwrite(path,face)
        cv.imshow('imaage',face)
        print(img_no)
        if cv.waitKey(1)=='q' or int(img_no)==n:
            break
    cv.destroyWindow('imaage')
    del cam


#To register with new face
def new_face():
    F_name=input('Enter your name \n')
    
    p_dir="R:/Face/model/new-train"               #images for train dataset
    path=os.path.join(p_dir, F_name)        
    os.makedirs(path)
    print("Images for train data set... Say cheese")
    capt_img(p_dir, F_name, 2000)

    p_dir="R:/Face/model/new-validation"               #images for validation dataset
    path=os.path.join(p_dir, F_name)        
    os.makedirs(path)
    print("Images for validation data set... Say cheese")
    capt_img(p_dir, F_name, 1400)
    
    model=train_model()
    return model,F_name

def attendance(name):
    #the directory is fixed
    os.chdir("R:/Face/model")
    #flag is assigned
    Flag=False
    #workbook is loaded(an external workbook was created earlier)
    wb=load_workbook('Attendance.xlsx')
    #worksheet is loaded
    ws=wb.active
    char1=get_column_letter(1)
    char2=get_column_letter(2)
    #checking for the same number and 40 Rs is deducted
    for i in range (2,5):
        if name==ws[char1+str(i)].value:
            ws[char2+str(i)]=(ws[char2+str(i)].value)+1
            total=ws[char2+str(i)].value
            Flag=True
            wb.save('Attendance.xlsx')
            break
    if Flag:
        detail=total
    else:
        detail='Student not found'
    return detail
    
Flag=True    
while True:
    n=int(input('Press 1 to add a new face and 2 to recognise the face and 3 to END \n'))
    if n==1:
        model,F_name=new_face()
        Flag=False
    elif n==2:
        if Flag:
            F_name=''
            model = load_model('FR1.h5')
            name=predict_face(model,F_name)
        else:
            name=predict_face(model,F_name)
    else:
        break
    total=attendance(name)
    print('Total attendance including today is ',total)
print('Thank You')           
        
